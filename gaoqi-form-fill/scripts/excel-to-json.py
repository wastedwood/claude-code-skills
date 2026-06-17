#!/usr/bin/env python3
"""
Convert a Gaoqi Excel workbook into the internal JSON data package.

This script is intentionally plain Python so Codex, Claude Code, and a normal
terminal can run the same workflow.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "缺少 Python 依赖 openpyxl。请先在当前 Python 环境安装 openpyxl 后再运行。"
    ) from exc


FIELD_ALIASES = {
    "RD号": ["RD号", "研发编号", "项目编号", "RD编号"],
    "研发项目(RD)名称": ["研发项目(RD)名称", "研发项目名称", "项目名称", "RD名称"],
    "IP号": ["IP号", "知识产权编号", "知识产权序号", "IP编号"],
    "知识产权名称": ["知识产权名称", "专利名称", "IP名称", "软著名称"],
    "专利号/登记号": ["专利号/登记号", "专利号", "登记号", "授权公告号"],
    "申请号": ["申请号", "专利申请号"],
    "名称": ["名称", "产品名称", "服务名称", "PS名称"],
    "上年度销售收入": ["上年度销售收入", "销售收入", "销售额"],
}


def norm_header(value: Any) -> str:
    return str(value or "").replace("\n", "").replace("\r", "").strip()


def build_header_map(ws) -> dict[str, int]:
    raw = {norm_header(cell.value): cell.column for cell in ws[1]}
    result = dict(raw)
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in raw:
                result[canonical] = raw[alias]
                break
    return result


def get_cell(row: tuple[Any, ...], headers: dict[str, int], name: str, default: Any = "") -> Any:
    col = headers.get(name)
    if not col:
        return default
    if col - 1 >= len(row):
        return default
    value = row[col - 1]
    if value is None:
        return default
    return value


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def as_date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return as_text(value)


def split_codes(value: Any) -> list[str]:
    text = as_text(value)
    if not text:
        return []
    parts = re.split(r"[、,，;/；\s]+", text)
    return [part.strip() for part in parts if part.strip()]


def normalize_ip_code(value: Any) -> str:
    text = as_text(value)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    if match:
        return f"IP{int(match.group(0)):02d}"
    return text.upper()


def normalize_rd_code(value: Any) -> str:
    text = as_text(value)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    if match:
        return f"RD{int(match.group(0)):02d}"
    return text.upper()


def normalize_ps_code(value: Any, fallback_index: int | None = None) -> str:
    text = as_text(value)
    if text.upper().startswith("PS"):
        match = re.search(r"PS\s*(\d+)", text, flags=re.I)
        if match:
            return f"PS{int(match.group(1)):02d}"
        return text.upper()
    match = re.search(r"\d+", text)
    if match:
        return f"PS{int(match.group(0)):02d}"
    if fallback_index is not None:
        return f"PS{fallback_index:02d}"
    return text


def patent_for_system(value: Any) -> str:
    text = as_text(value)
    if text.startswith("ZL"):
        return text
    if text.startswith("CN"):
        return "ZL" + text[2:]
    return text


def iter_records(wb, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
    if sheet_name not in wb.sheetnames:
        return [], {}
    ws = wb[sheet_name]
    headers = build_header_map(ws)
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        item = {}
        for header, col in headers.items():
            if col - 1 < len(row):
                item[header] = row[col - 1]
        records.append(item)
    return records, headers


def parse_ip(wb) -> list[dict[str, Any]]:
    records, _ = iter_records(wb, "IP明细")
    items = []
    for idx, row in enumerate(records, start=1):
        code = normalize_ip_code(row.get("IP号")) or f"IP{idx:02d}"
        items.append(
            {
                "no": re.sub(r"\D", "", code) or f"{idx:02d}",
                "code": code,
                "name": as_text(row.get("知识产权名称")),
                "category": as_text(row.get("类别")),
                "categoryValue": "",
                "accreditNo": patent_for_system(row.get("专利号/登记号")),
                "applicationNo": as_text(row.get("申请号")),
                "authorizedDate": as_date_text(row.get("授权日期")),
                "acquireMethod": as_text(row.get("获得方式")),
                "acquireMethodValue": "",
                "source": "Excel",
                "usableForSystemValidation": False,
                "notes": "示例或待用户确认数据",
            }
        )
    return items


def parse_human_resources(wb) -> dict[str, Any]:
    records, _ = iter_records(wb, "人员花名册")
    total = len(records)
    tech_total = sum(1 for row in records if "是" in as_text(row.get("是否研发人员")))

    education = {"doctor": 0, "master": 0, "bachelor": 0, "juniorCollegeOrBelow": 0}
    title = {"senior": 0, "middle": 0, "primary": 0, "seniorTechnician": 0}
    for row in records:
        edu = as_text(row.get("学历"))
        if "博士" in edu:
            education["doctor"] += 1
        elif "硕士" in edu or "研究生" in edu:
            education["master"] += 1
        elif "本科" in edu or "学士" in edu:
            education["bachelor"] += 1
        else:
            education["juniorCollegeOrBelow"] += 1

        job_title = as_text(row.get("职称"))
        if "高级技师" in job_title:
            title["seniorTechnician"] += 1
        elif "高级" in job_title:
            title["senior"] += 1
        elif "中级" in job_title:
            title["middle"] += 1
        elif job_title:
            title["primary"] += 1

    return {
        "employeeTotal": total,
        "techStaffTotal": tech_total,
        "employment": {
            "onJob": {"employee": total, "tech": tech_total},
            "partTime": {"employee": 0, "tech": 0},
            "temporary": {"employee": 0, "tech": 0},
            "foreign": {"employee": 0, "tech": 0},
            "returnedOverseas": {"employee": 0, "tech": 0},
            "talentPlan": {"employee": 0, "tech": 0},
        },
        "education": education,
        "title": title,
        "age": {"under30": 0, "age31To40": 0, "age41To50": 0, "over51": 0},
    }


def parse_rd_projects(wb) -> list[dict[str, Any]]:
    records, _ = iter_records(wb, "RD-IP-PS、RD文本")
    projects = []
    for idx, row in enumerate(records, start=1):
        code = normalize_rd_code(row.get("RD号")) or f"RD{idx:02d}"
        related_ips = [normalize_ip_code(code) for code in split_codes(row.get("对应IP序号"))]
        related_ips = [code for code in related_ips if code]
        projects.append(
            {
                "code": code,
                "no": re.sub(r"\D", "", code) or f"{idx:02d}",
                "name": as_text(row.get("研发项目(RD)名称")),
                "startDate": as_date_text(row.get("开始日期")),
                "endDate": as_date_text(row.get("结束日期")),
                "domain": {"level1": "", "level1Value": "", "level2": "", "level2Value": "", "level3": "", "level3Value": ""},
                "technologySource": "企业自有技术",
                "technologySourceValue": "",
                "relatedIpCodes": related_ips,
                "budgetTotal": as_number(row.get("研发预算")),
                "spendingByYear": {
                    "2023": as_number(row.get("2023研发费")),
                    "2024": as_number(row.get("2024研发费")),
                    "2025": as_number(row.get("2025研发费")),
                },
                "purposeAndOrganization": as_text(row.get("研发目的组织方式")),
                "coreTechnologyAndInnovation": as_text(row.get("核心技术及创新点")),
                "achievements": as_text(row.get("阶段性成果")),
                "orderNo": idx,
            }
        )
    return projects


def parse_rd_fees(wb) -> list[dict[str, Any]]:
    items = []
    for year in (2023, 2024, 2025):
        sheet = f"{year}年研发费用结构明细表"
        records, _ = iter_records(wb, sheet)
        for idx, row in enumerate(records, start=1):
            rd_code = normalize_rd_code(row.get("项目编号"))
            if not rd_code:
                continue
            items.append(
                {
                    "rdCode": rd_code,
                    "year": year,
                    "internal": {
                        "personnel": as_number(row.get("人员人工费用")),
                        "directInput": as_number(row.get("直接投入费用")),
                        "depreciation": as_number(row.get("折旧摊销")),
                        "amortization": as_number(row.get("无形资产摊销")),
                        "design": as_number(row.get("设计费")),
                        "equipmentDebugTrial": as_number(row.get("装备调试费用与试验费用")),
                        "other": as_number(row.get("其他费用")),
                    },
                    "external": {
                        "entrusted": as_number(row.get("委外研发费用")),
                        "domesticEntrusted": as_number(row.get("境内的外部研发费用")),
                    },
                    "fillUser": "测试填报人",
                    "fillDate": date.today().isoformat(),
                    "orderNo": idx,
                }
            )
    return items


def parse_products(wb) -> list[dict[str, Any]]:
    records, _ = iter_records(wb, "PS文本")
    products = []
    for idx, row in enumerate(records, start=1):
        code = normalize_ps_code(row.get("序号"), idx)
        related_ips = [normalize_ip_code(code) for code in split_codes(row.get("对应知识产权编号"))]
        related_ips = [code for code in related_ips if code]
        products.append(
            {
                "code": code,
                "no": re.sub(r"\D", "", code) or f"{idx:02d}",
                "name": as_text(row.get("名称")),
                "domain": {"level1": "", "level1Value": "", "level2": "", "level2Value": "", "level3": "", "level3Value": ""},
                "technologySource": "企业自有技术",
                "technologySourceValue": "",
                "salesRevenue": as_number(row.get("上年度销售收入")),
                "isMainProduct": False,
                "relatedIpCodes": related_ips,
                "keyTechnology": as_text(row.get("关键技术")),
                "mainTechnicalIndicators": as_text(row.get("主要技术指标")),
                "competitiveAdvantage": as_text(row.get("竞争优势")),
                "ipSupport": as_text(row.get("知识产权详情及支持作用")),
                "orderNo": idx,
            }
        )
    return products


def parse_innovation(wb) -> dict[str, Any]:
    summary_records, _ = iter_records(wb, "企业创新能力")
    summaries = {as_text(row.get("主题")): as_text(row.get("内容摘要")) for row in summary_records}

    transform_records, _ = iter_records(wb, "科技成果转化")
    transformations = []
    for idx, row in enumerate(transform_records, start=1):
        transformations.append(
            {
                "name": as_text(row.get("科技成果名称")),
                "type": as_text(row.get("成果类型")),
                "typeValue": "",
                "source": as_text(row.get("科技成果来源")),
                "sourceValue": "",
                "result": as_text(row.get("转化结果")),
                "resultValue": "",
                "year": as_text(row.get("转化时间")),
                "relatedIpCodes": [normalize_ip_code(code) for code in split_codes(row.get("关联IP")) if normalize_ip_code(code)],
                "relatedRdCodes": [normalize_rd_code(code) for code in split_codes(row.get("关联RD")) if normalize_rd_code(code)],
                "relatedPsCodes": [normalize_ps_code(code) for code in split_codes(row.get("对应PS")) if normalize_ps_code(code)],
                "ways": [],
                "wayValues": [],
                "orderNo": idx,
                "notes": as_text(row.get("说明")),
            }
        )

    return {
        "ipCompetitiveness": summaries.get("知识产权", ""),
        "transformationSummary": summaries.get("成果转化", ""),
        "transformations": transformations,
        "rdOrganizationManagement": summaries.get("研发组织管理", ""),
        "managementAndTechStaff": summaries.get("科技人员", ""),
    }


def convert(input_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    return {
        "meta": {
            "companyName": "测试企业有限公司",
            "reportYear": date.today().year,
            "dataInnocomId": "",
            "sourceFiles": [input_path.name],
            "notes": "由 Excel 输入适配层生成",
        },
        "intellectualProperties": parse_ip(wb),
        "humanResources": parse_human_resources(wb),
        "rdProjects": parse_rd_projects(wb),
        "rdFees": parse_rd_fees(wb),
        "products": parse_products(wb),
        "innovation": parse_innovation(wb),
        "standards": [],
        "businessSummary": {
            "netAssets": {"2023": 0, "2024": 0, "2025": 0},
            "salesRevenue": {"2023": 0, "2024": 0, "2025": 0},
            "profit": {"2023": 0, "2024": 0, "2025": 0},
            "domesticRdExpense": {"2023": 0, "2024": 0, "2025": 0},
            "totalRevenue": 0,
        },
    }


def main() -> int:
    if len(sys.argv) != 3:
        print("用法: python gaoqi-form-fill/scripts/excel-to-json.py 输入.xlsx 输出.json", file=sys.stderr)
        return 2

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    if not input_path.exists():
        print(f"找不到输入文件: {input_path}", file=sys.stderr)
        return 2

    data = convert(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"已生成: {output_path}")
    print(
        "统计: "
        f"IP {len(data['intellectualProperties'])} 条, "
        f"RD {len(data['rdProjects'])} 条, "
        f"费用 {len(data['rdFees'])} 条, "
        f"产品 {len(data['products'])} 条, "
        f"成果转化 {len(data['innovation']['transformations'])} 条"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
